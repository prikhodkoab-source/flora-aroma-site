from __future__ import annotations

import csv
import shutil
import subprocess
import tempfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PRODUCTS = ROOT / "data" / "products.csv"
LOGO = ROOT / "public" / "images" / "logo.png"
OUTPUT_DIR = ROOT / "public" / "downloads"
OUTPUT_PDF = OUTPUT_DIR / "flora-aroma-price.pdf"
SOFFICE_CANDIDATES = (
    Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
    Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
)


def split_values(value: str) -> list[str]:
    return [item.strip() for item in value.split(";") if item.strip()]


def read_price_rows() -> list[dict[str, str]]:
    with PRODUCTS.open("r", encoding="utf-8-sig", newline="") as handle:
        products = list(csv.DictReader(handle))

    rows: list[dict[str, str]] = []
    for product in products:
        containers = split_values(product.get("variant_containers", ""))
        prices = split_values(product.get("variant_prices_uah", ""))
        units = split_values(product.get("variant_units", ""))
        if not containers:
            containers = [product["container"]]
            prices = [product["price_uah"]]
            units = [product["unit"]]

        if not (len(containers) == len(prices) == len(units)):
            raise RuntimeError(f"variant_mismatch:{product['plant_id']}")

        for container, price, unit in zip(containers, prices, units):
            rows.append(
                {
                    "plant_id": product["plant_id"],
                    "name_uk": product["name_uk"],
                    "latin_name": product["latin_name"],
                    "container": container,
                    "price": f"{price} UAH/{unit}",
                }
            )

    return sorted(rows, key=lambda row: (row["name_uk"], row["container"]))


def build_workbook(rows: list[dict[str, str]], target: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Прайс"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:E2")
    ws["B1"] = "FLORA & AROMA"
    ws["B1"].font = Font(name="Arial", size=26, bold=True, color="1F5D42")
    ws["B1"].alignment = Alignment(vertical="center")
    ws.merge_cells("B3:E3")
    ws["B3"] = "Актуальний асортимент рослин"
    ws["B3"].font = Font(name="Arial", size=15, bold=True, color="18352B")
    ws.merge_cells("B4:E4")
    ws["B4"] = f"Дата формування: {date.today():%d.%m.%Y}  |  flora-aroma.com.ua  |  +380500272882"
    ws["B4"].font = Font(name="Arial", size=9, color="66736F")

    if LOGO.is_file():
        logo = Image(LOGO)
        logo.width = 92
        logo.height = 92
        ws.add_image(logo, "A1")

    headers = ("plant_id", "Українська назва", "Латинська назва", "Горщик / касета", "Ціна")
    header_row = 6
    for column, header in enumerate(headers, start=1):
        cell = ws.cell(header_row, column, header)
        cell.font = Font(name="Arial", size=9, bold=True, color="18352B")
        cell.fill = PatternFill("solid", fgColor="DDEEDF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="D8E4DC")
    for row_index, row in enumerate(rows, start=header_row + 1):
        values = (
            row["plant_id"],
            row["name_uk"],
            row["latin_name"],
            row["container"],
            row["price"],
        )
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row_index, column, value)
            cell.font = Font(name="Arial", size=8)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

    widths = (15, 31, 29, 25, 15)
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[6].height = 28
    ws.freeze_panes = "A7"
    ws.auto_filter.ref = f"A6:E{header_row + len(rows)}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.oddFooter.center.text = "Наявність і можливість резерву підтверджує оператор."
    ws.oddFooter.center.size = 8
    ws.oddFooter.center.color = "66736F"
    wb.save(target)


def find_soffice() -> Path:
    configured = shutil.which("soffice")
    if configured:
        return Path(configured)
    for candidate in SOFFICE_CANDIDATES:
        if candidate.is_file():
            return candidate
    raise RuntimeError("LibreOffice soffice is required to generate the PDF")


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = read_price_rows()
    with tempfile.TemporaryDirectory(prefix="flora-price-") as temp_dir:
        temp = Path(temp_dir)
        workbook = temp / "flora-aroma-price.xlsx"
        libreoffice_profile = temp / "libreoffice-profile"
        libreoffice_profile.mkdir()
        build_workbook(rows, workbook)
        subprocess.run(
            [
                str(find_soffice()),
                "--headless",
                f"-env:UserInstallation={libreoffice_profile.as_uri()}",
                "--convert-to",
                "pdf",
                "--outdir",
                str(temp),
                str(workbook),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        generated = temp / "flora-aroma-price.pdf"
        if not generated.is_file():
            raise RuntimeError("LibreOffice did not produce the expected PDF")
        shutil.copy2(generated, OUTPUT_PDF)

    print(f"price_rows={len(rows)}")
    print(f"output={OUTPUT_PDF.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
